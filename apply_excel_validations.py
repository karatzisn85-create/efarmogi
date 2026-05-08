#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script to copy data validations from template Excel to exported Excel
"""

import sys
import os
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
import shutil

# Fix Windows console encoding
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

def apply_validations(template_path, target_path):
    """
    Copy data validations from template to target Excel file
    """
    try:
        print(f"📂 Loading template: {template_path}")
        template_wb = load_workbook(template_path, data_only=False)
        
        print(f"📂 Loading target: {target_path}")
        target_wb = load_workbook(target_path, data_only=False)
        
        template_sheet_name = '02. ΠΙΝΑΚΑΣ ΕΡΓΩΝ'
        
        if template_sheet_name not in template_wb.sheetnames:
            print(f"❌ Sheet '{template_sheet_name}' not found in template")
            return False
            
        if template_sheet_name not in target_wb.sheetnames:
            print(f"❌ Sheet '{template_sheet_name}' not found in target")
            return False
        
        template_sheet = template_wb[template_sheet_name]
        target_sheet = target_wb[template_sheet_name]
        
        # Get number of rows in target
        target_rows = target_sheet.max_row
        print(f"📊 Target has {target_rows} rows")
        
        # Copy data validations from template
        validations_copied = 0
        
        for dv in template_sheet.data_validations.dataValidation:
            print(f"🔍 Found validation: {dv.sqref} - Type: {dv.type}")
            
            # Create new validation with same properties
            new_dv = DataValidation(
                type=dv.type,
                formula1=dv.formula1,
                formula2=dv.formula2,
                allow_blank=dv.allow_blank,
                showDropDown=dv.showDropDown,
                showErrorMessage=dv.showErrorMessage,
                showInputMessage=dv.showInputMessage,
                errorTitle=dv.errorTitle,
                error=dv.error,
                promptTitle=dv.promptTitle,
                prompt=dv.prompt
            )
            
            # Apply to all data rows in target (from row 2 to max_row)
            # Get the column letter from the original validation
            original_ref = str(dv.sqref)
            
            # Parse column from reference (e.g., "E2:E100" -> "E")
            if ':' in original_ref:
                col_start = original_ref.split(':')[0].rstrip('0123456789')
                col_end = original_ref.split(':')[1].rstrip('0123456789')
            else:
                col_start = original_ref.rstrip('0123456789')
                col_end = col_start
            
            # Apply validation to all rows in target
            if col_start == col_end:
                new_range = f"{col_start}2:{col_start}{target_rows}"
            else:
                new_range = f"{col_start}2:{col_end}{target_rows}"
            
            new_dv.sqref = new_range
            target_sheet.add_data_validation(new_dv)
            validations_copied += 1
            print(f"✅ Applied validation to range: {new_range}")
        
        # Save target workbook
        print(f"💾 Saving target file...")
        target_wb.save(target_path)
        
        print(f"✅ Successfully copied {validations_copied} validations")
        return True
        
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python apply_excel_validations.py <template_path> <target_path>")
        sys.exit(1)
    
    template_path = sys.argv[1]
    target_path = sys.argv[2]
    
    if not os.path.exists(template_path):
        print(f"❌ Template file not found: {template_path}")
        sys.exit(1)
    
    if not os.path.exists(target_path):
        print(f"❌ Target file not found: {target_path}")
        sys.exit(1)
    
    success = apply_validations(template_path, target_path)
    sys.exit(0 if success else 1)

