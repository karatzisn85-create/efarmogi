#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Quick verification script for exported Excel
"""

import sys
import os
from openpyxl import load_workbook

# Fix Windows console encoding
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

def verify_excel(file_path):
    """Verify exported Excel file"""
    try:
        print(f"\n{'='*70}")
        print(f"📊 ΕΛΕΓΧΟΣ ΕΞΑΓΟΜΕΝΟΥ EXCEL")
        print(f"{'='*70}\n")
        
        wb = load_workbook(file_path, data_only=False)
        sheet = wb['02. ΠΙΝΑΚΑΣ ΕΡΓΩΝ']
        
        # 1. Check data validations
        print("1️⃣  DATA VALIDATIONS:")
        validation_count = len(sheet.data_validations.dataValidation)
        print(f"   ✅ Βρέθηκαν {validation_count} validations")
        
        for dv in sheet.data_validations.dataValidation:
            print(f"      - {dv.sqref} (Type: {dv.type})")
        
        # 2. Check first data row
        print("\n2️⃣  ΠΡΩΤΗ ΓΡΑΜΜΗ ΔΕΔΟΜΕΝΩΝ:")
        if sheet.max_row >= 2:
            row = 2
            ka = sheet[f'B{row}'].value
            title = sheet[f'C{row}'].value
            amount = sheet[f'L{row}'].value
            contract = sheet[f'P{row}'].value
            
            print(f"   ΚΑ: {ka}")
            print(f"   Τίτλος: {title}")
            print(f"   Προϋπολογισμός: {amount}")
            print(f"   Συμβατικό Αντικείμενο: {contract}")
        
        # 3. Check amounts format
        print("\n3️⃣  ΕΛΕΓΧΟΣ ΜΟΡΦΗΣ ΠΟΣΩΝ:")
        for row in range(2, min(6, sheet.max_row + 1)):
            amount = sheet[f'L{row}'].value
            if amount:
                print(f"   Γραμμή {row}: {amount}")
        
        # 4. Check sorting
        print("\n4️⃣  ΕΛΕΓΧΟΣ ΤΑΞΙΝΟΜΗΣΗΣ (ΚΑ):")
        ka_codes = []
        for row in range(2, min(7, sheet.max_row + 1)):
            ka = sheet[f'B{row}'].value
            if ka:
                ka_codes.append(ka)
        print(f"   Πρώτοι 5 ΚΑ: {', '.join(ka_codes)}")
        
        # 5. General info
        print("\n5️⃣  ΓΕΝΙΚΑ ΣΤΟΙΧΕΙΑ:")
        print(f"   Συνολικές γραμμές: {sheet.max_row - 1}")
        print(f"   Στήλες: {sheet.max_column}")
        
        print(f"\n{'='*70}")
        print("✅ ΕΛΕΓΧΟΣ ΟΛΟΚΛΗΡΩΘΗΚΕ")
        print(f"{'='*70}\n")
        
        return True
        
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python verify_export.py <excel_file>")
        sys.exit(1)
    
    file_path = sys.argv[1]
    
    if not os.path.exists(file_path):
        print(f"❌ File not found: {file_path}")
        sys.exit(1)
    
    success = verify_excel(file_path)
    sys.exit(0 if success else 1)

